# -*- coding:utf-8 -*-
import io
import os
from os import rename, listdir
import sys

# reader .csv file
import csv

# http caller
import requests

# create xls, xlsx
import openpyxl
from openpyxl import Workbook

# check system info
import platform
from time import sleep
from datetime import datetime

# Crawling lib
from bs4 import BeautifulSoup

# Multi Processing lib
import multiprocessing
from multiprocessing import Process

# argparse
import argparse

# email sender
from _email import send as _email

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

_format = 1
# _format 0 : .csv
# _format 1 : .xlsx
# _format 2 : .xls

c_type = False
g_type = False
d_type = False
r_type = False

c_rank = 1
g_rank = 1
d_rank = 1
r_rank = 1

# 임시 파일 제거 함수
def file_delete(find):
    list = os.listdir(os.getcwd()+"/tmp")
    
    if find == "cpu":
        for i in list:
            if i.find("cpu.csv") != -1 :
                os.remove("tmp/" + i)
    elif find == "gpu":
        for i in list:
            if i.find("gpu.csv") != -1 :
                os.remove("tmp/" + i)
    elif find == "drive":
        for i in list:
            if i.find("drive.csv") != -1 :
                os.remove("tmp/" + i)
    elif find == "ram":
        for i in list:
            if i.find("ram.csv") != -1 :
                os.remove("tmp/" + i)

def dayfilename(find):
    now = datetime.now()
    day = '%s-%s-%s' % (now.year,now.month,now.day)
    files = listdir('.')
    
    if find == "cpu":
        for name in files:
            if "cpu.csv" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "cpu.xlsx" in name:
                if len(name) < 9:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "cpu.xls" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)
    elif find == "gpu":
        for name in files:
            if "gpu.csv" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "gpu.xlsx" in name:
                if len(name) < 9:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "gpu.xls" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)
    elif find == "drive":
        for name in files:
            if "drive.csv" in name:
                if len(name) < 12:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "drive.xlsx" in name:
                if len(name) < 13:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "drive.xls" in name:
                if len(name) < 12:
                    new_name = day + "-"+ name
                    rename(name,new_name)
    elif find == "ram":
        for name in files:
            if "ram.csv" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "ram.xlsx" in name:
                if len(name) < 9:
                    new_name = day + "-"+ name
                    rename(name,new_name)
            elif "ram.xls" in name:
                if len(name) < 8:
                    new_name = day + "-"+ name
                    rename(name,new_name)


def convert_extention(find):
    """
    파일 확장자 선택 함수
    """
    filename = ''+find+'.xlsx'
    _filename = ''+find+'.csv'
    
    xlsx = openpyxl.load_workbook(filename)
    csv = open(_filename, "w+", -1, "UTF8")
    for sheet_name in xlsx:
        sheet = sheet_name
        data = sheet.rows
        for row in data:
            l = list(row)
            for i in range(len(l)):
                if i == len(l) - 1:
                    csv.write(str(l[i].value))
                    csv.write('\n')
                else:
                    csv.write(str(l[i].value) + ',')
    csv.close()
    os.remove(filename)

def convert_excel(find):
    """
    xlsx 파일 변환기
    """
    global _format
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    
    
    if find == "cpu":
        ws1.title = '최상위 CPU'
        ws2.title = '중상위 CPU'
        ws3.title = '중하위 CPU'
        ws4.title = '하위 CPU'
    elif find == "gpu":
        ws1.title = '최상위 GPU'
        ws2.title = '중상위 GPU'
        ws3.title = '중하위 GPU'
        ws4.title = '하위 GPU'
    elif find == "drive":
        ws1.title = '최상위 Drive'
        ws2.title = '중상위 Drive'
        ws3.title = '중하위 Drive'
        ws4.title = '하위 Drive'
    elif find == "ram":
        ws5 = wb.create_sheet()
        ws6 = wb.create_sheet()
        ws1.title = 'DDR4 RAM Read'
        ws2.title = 'DDR4 RAM Write'
        ws3.title = 'DDR4 RAM Latency'
        ws4.title = 'DDR3 RAM Read'
        ws5.title = 'DDR3 RAM Write'
        ws6.title = 'DDR3 RAM Latency'
    
    if _format == 0:
        savefile = "" + find + ".xlsx"
    elif _format == 1:
        savefile = "" + find + ".xlsx"
    elif _format == 2:
        savefile = "" + find + ".xls"

    CSV_SEPARATOR = ","

    try:
        column = 'B'
        if find == "cpu":
            # 각각을 쓰레드로 실행해볼까?
            with open("tmp/1_cpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws1.cell(r+1,c+1,val)

            with open("tmp/2_cpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws2.cell(r+1,c+1,val)

            with open("tmp/3_cpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws3.cell(r+1,c+1,val)

            with open("tmp/4_cpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws4.cell(r+1,c+1,val)

        elif find == "gpu":
            with open("tmp/1_gpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws1.cell(r+1,c+1,val)

            with open("tmp/2_gpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws2.cell(r+1,c+1,val)

            with open("tmp/3_gpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws3.cell(r+1,c+1,val)

            with open("tmp/4_gpu.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws4.cell(r+1,c+1,val)

        elif find == "drive":
            with open("tmp/1_drive.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws1.cell(r+1,c+1,val)

            with open("tmp/2_drive.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws2.cell(r+1,c+1,val)

            with open("tmp/3_drive.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws3.cell(r+1,c+1,val)

            with open("tmp/4_drive.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws4.cell(r+1,c+1,val)
        elif find == "ram":
            with open("tmp/r_ddr4_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws1.cell(r+1,c+1,val)

            with open("tmp/w_ddr4_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws2.cell(r+1,c+1,val)

            with open("tmp/l_ddr4_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws3.cell(r+1,c+1,val)

            with open("tmp/r_ddr3_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws4.cell(r+1,c+1,val)

            with open("tmp/w_ddr3_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws5.cell(r+1,c+1,val)
                            ws5.column_dimensions[column].width = 40

            with open("tmp/l_ddr3_ram.csv", encoding='UTF8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                            ws6.cell(r+1,c+1,val)
                            ws6.column_dimensions[column].width = 40
        ws1.column_dimensions[column].width = 40
        ws2.column_dimensions[column].width = 40
        ws3.column_dimensions[column].width = 40
        ws4.column_dimensions[column].width = 40
    except:
        return
    wb.save(savefile)
    
    if _format == 0:
        convert_extention(find)
        
def crawling_thread(id, _type, web, flag_ram=""):
    """
    크롤링 쓰레드
    """
    res = requests.get(web)
    soup = BeautifulSoup(res.content, 'lxml')
    try:
        data = soup.find("ul",{"class": "chartlist"}).get_text().split("*")
    except AttributeError:
        _email.send()
        print("\n죄송합니다. 원본사이트에 접근 할 수 없습니다.\n확인 후 다음 업데이트에 적용하겠습니다.")
        return
    if _type == "cpu":
        _file = str(id)+"_"+_type
        f = open("tmp/"+_file+".csv", 'w+', encoding='UTF8')
        for i in data:
            f.write(i)
        f.close()
        make_csv_new(_file)
    elif _type == "gpu":
        _file = str(id)+"_"+_type 
        f = open("tmp/"+_file+".csv", 'w+', encoding='UTF8')
        for i in data:
            f.write(i)
        f.close()
        make_csv_new_g(_file)
    elif _type == "drive":
        _file = str(id)+"_"+_type 
        f = open("tmp/"+_file+".csv", 'w+', encoding='UTF8')
        for i in data:
            f.write(i)
        f.close()
        make_csv_new_d(_file)
    elif _type == "ram":
        if id < 4:
            _file = flag_ram+"_ddr4_"+_type
        else:
            _file = flag_ram+"_ddr3_"+_type
        f = open("tmp/"+_file+".csv", 'w+', encoding='UTF8')
        for i in data:
            f.write(i)
        f.close()
        make_csv_new_r(_file)
    # print(str(id) + ": Finish")
    return

# CPU Crawling
def extract_cpu():
    try:
        th1 = Process(target=crawling_thread, args=(1, "cpu", 'https://www.cpubenchmark.net/high_end_cpus.html'))
        th2 = Process(target=crawling_thread, args=(2, "cpu", 'https://www.cpubenchmark.net/mid_range_cpus.html'))
        th3 = Process(target=crawling_thread, args=(3, "cpu", 'https://www.cpubenchmark.net/midlow_range_cpus.html'))
        th4 = Process(target=crawling_thread, args=(4, "cpu", 'https://www.cpubenchmark.net/low_end_cpus.html'))
        th1.start(),th2.start(),th3.start(),th4.start()
        th1.join(),th2.join(),th3.join(),th4.join()
    except:
        return
    print('CPU Data Extract Complete!!!')
    convert_excel("cpu")
    file_delete("cpu")
    dayfilename("cpu")

# GPU Crawling
def extract_gpu():
    try:
        th1 = Process(target=crawling_thread, args=(1, "gpu", 'https://www.videocardbenchmark.net/high_end_gpus.html'))
        th2 = Process(target=crawling_thread, args=(2, "gpu", 'https://www.videocardbenchmark.net/mid_range_gpus.html'))
        th3 = Process(target=crawling_thread, args=(3, "gpu", 'https://www.videocardbenchmark.net/midlow_range_gpus.html'))
        th4 = Process(target=crawling_thread, args=(4, "gpu", 'https://www.videocardbenchmark.net/low_end_gpus.html'))
        th1.start(),th2.start(),th3.start(),th4.start()
        th1.join(),th2.join(),th3.join(),th4.join()
    except:
        return
    print('GPU Data Extract Complete!!!')
    convert_excel("gpu")
    file_delete("gpu")
    dayfilename("gpu")
    
# Drive Crawling
def extract_drive():
    try:
        th1 = Process(target=crawling_thread, args=(1, "drive", 'https://www.harddrivebenchmark.net/high_end_drives.html'))
        th2 = Process(target=crawling_thread, args=(2, "drive", 'https://www.harddrivebenchmark.net/mid_range_drives.html'))
        th3 = Process(target=crawling_thread, args=(3, "drive", 'https://www.harddrivebenchmark.net/low_mid_range_drives.html'))
        th4 = Process(target=crawling_thread, args=(4, "drive", 'https://www.harddrivebenchmark.net/low_end_drives.html'))
        th1.start(),th2.start(),th3.start(),th4.start()
        th1.join(),th2.join(),th3.join(),th4.join()
    except:
        return
    print('Drive Data Extract Complete!!!')
    convert_excel("drive")
    file_delete("drive")
    dayfilename("drive")

# RAM Crawling
def extract_ram():
    try:
        th1 = Process(target=crawling_thread, args=(1, "ram", 'https://www.memorybenchmark.net/read_uncached_ddr4_intel.html', "r"))
        th2 = Process(target=crawling_thread, args=(2, "ram", 'https://www.memorybenchmark.net/write_ddr4_intel.html', "w"))
        th3 = Process(target=crawling_thread, args=(3, "ram", 'https://www.memorybenchmark.net/latency_ddr4_intel.html', "l"))
        th4 = Process(target=crawling_thread, args=(4, "ram", 'https://www.memorybenchmark.net/read_uncached_ddr3_intel.html', "r"))
        th5 = Process(target=crawling_thread, args=(5, "ram", 'https://www.memorybenchmark.net/write_ddr3_intel.html', "w"))
        th6 = Process(target=crawling_thread, args=(6, "ram", 'https://www.memorybenchmark.net/latency_ddr3_intel.html', "l"))
        th1.start(),th2.start(),th3.start(),th4.start(),th5.start(),th6.start()
        th1.join(),th2.join(),th3.join(),th4.join(),th5.join(),th6.join()
    except:
        return
    print('RAM Data Extract Complete!!!')
    convert_excel("ram")
    file_delete("ram")
    dayfilename("ram")

# Make CPU Data
def make_csv_new(name):
    str = []
    tmp = []
    test = ""
    global c_rank

    f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
    lines = f.readlines()
    for line in lines[1:]:
        str.append(line)
    f.close()

    for i in str:
        i = i.replace(",","").strip()
        if "NA" in i:
            i = i[:i.find("NA")]
            if i[i.find('%') - 2] != '(':#두자리 수
                str_rank = repr(c_rank)
                tmp.append(str_rank + "," + i[:i.index('%') - 3] + "," + i[i.index('%') + 2:])
                c_rank+=1
            elif i[i.find('%') - 2] == '(':#한자리 수
                str_rank = repr(c_rank)
                tmp.append(str_rank + "," + i[:i.index('%') - 2] + "," + i[i.index('%') + 2:])
                c_rank+=1
        elif "$" in i:
            i = i[:i.find('$')]
            if i[i.find('%') - 2] != '(':#두자리 수
                str_rank = repr(c_rank)
                tmp.append(str_rank + "," + i[:i.index('%') - 3] + "," + i[i.index('%') + 2:])
                c_rank+=1
            elif i[i.find('%') - 2] == '(':#한자리 수
                str_rank = repr(c_rank)
                tmp.append(str_rank + "," + i[:i.index('%') - 2] + "," + i[i.index('%') + 2:])
                c_rank+=1
                

    f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
    for i in tmp:
        f.write(i)
        f.write("\n")

    f.close()

# Make GPU Data
def make_csv_new_g(name):
    str = []
    tmp = []
    count = 1
    global g_rank
    
    f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
    lines = f.readlines()
    for line in lines[:]:
        str.append(line)
    f.close()
    
    for i in str:
        if len(i) > 1:
            i = i.replace(",","").strip()
            if count%3 == 1:
                str_rank = repr(g_rank)
                tmp.append(str_rank + "," + i)
                g_rank+=1
                count+=1
            elif count%3 == 2:
                tmp.append(i[i.find(')') + 2:])
                count+=1
            elif count%3 == 0:
                count+=1

    count = 1

    f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
    for i in tmp:
        if count%2 == 0:
            f.write(i)
            f.write("\n")
            count+=1
        else:
            f.write(i + ",")
            count+=1
    f.close()

# Make Drive Data
def make_csv_new_d(name):
    str = []
    tmp = []
    count = 1
    global d_rank
    
    f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
    lines = f.readlines()
    for line in lines[:]:
        str.append(line)
    f.close()
    
    for i in str:
        if len(i) > 1:
            i = i.replace(",","").strip()
            if count%3 == 1:
                str_rank = repr(d_rank)
                tmp.append(str_rank + "," + i)
                d_rank+=1
                count+=1
            elif count%3 == 2:
                tmp.append(i[i.find(')') + 2:])
                count+=1
            elif count%3 == 0:
                count+=1

    count = 1

    f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
    for i in tmp:
        if count%2 == 0:
            f.write(i)
            f.write("\n")
            count+=1
        else:
            f.write(i + ",")
            count+=1
    f.close()

# Make RAM Data
def make_csv_new_r(name):
    str = []
    tmp = []
    count = 1
    global r_rank
    
    f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
    lines = f.readlines()
    for line in lines[:]:
        str.append(line)
    f.close()
    
    for i in str:
        if len(i) > 1:
            i = i.replace(",","").strip()
            if count%3 == 1:
                str_rank = repr(r_rank)
                tmp.append(str_rank + "," + i)
                r_rank+=1
                count+=1
            elif count%3 == 2:
                tmp.append(i[i.find(')') + 2:])
                count+=1
            elif count%3 == 0:
                count+=1

    count = 1

    f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
    for i in tmp:
        if count%2 == 0:
            f.write(i)
            f.write("\n")
            count+=1
        else:
            f.write(i + ",")
            count+=1
    f.close()

# 시작 지점
if __name__ == "__main__":
    """
    windows system multi processing support code
    """
    if sys.platform.startswith('win'):
        multiprocessing.freeze_support()

    parser = argparse.ArgumentParser(description="AutoBench 사용설명서")
    parser.add_argument("-v", dest='version', action='store_const', const='1.1.3', help='버전 확인')
    parser.add_argument('-f', dest='format', metavar='F', type=str, nargs='+', help='포맷 형식 지정. (단, 하나만 선택 가능)', default='xlsx')
    parser.add_argument('-o', dest='option', metavar='O', type=str, nargs='+', help='추출물 지정', default=['cpu', 'gpu', 'drive', 'ram'])
    
    args = parser.parse_args()

    if args.version != None:
        print(args.version)
        sys.exit()

    if 'cpu' in args.option:
        c_type = True
    if 'gpu' in args.option:
        g_type = True
    if 'drive' in args.option:
        d_type = True
    if 'ram' in args.option:
        r_type = True
    
    if len(args.format) != 1 and type(args.format) != type(""):
        parser.print_help()
        sys.exit()

    if not os.path.exists("tmp"):
        os.mkdir("tmp")

    if 'csv' in args.format:
        _format = 0
    elif 'xlsx' in args.format:
        _format = 1
    elif 'xls' in args.format:
        _format = 2
    else:
        print("error")

    th_cpu = None
    th_gpu = None
    th_drive = None
    th_ram = None

    if c_type:
        th_cpu = Process(target=extract_cpu)
    if g_type:
        th_gpu = Process(target=extract_gpu)
    if d_type:
        th_drive = Process(target=extract_drive)
    if r_type:
        th_ram = Process(target=extract_ram)

    if th_cpu != None and th_gpu != None and th_drive != None and th_ram != None:
        th_cpu.start(), th_gpu.start(), th_drive.start(), th_ram.start()
        th_cpu.join(),th_gpu.join(),th_drive.join(),th_ram.join()
    elif th_cpu == None and th_gpu != None and th_drive != None and th_ram != None:
        th_gpu.start(), th_drive.start(), th_ram.start()
        th_gpu.join(),th_drive.join(),th_ram.join()
    elif th_cpu != None and th_gpu == None and th_drive != None and th_ram != None:
        th_cpu.start(), th_drive.start(), th_ram.start()
        th_cpu.join(), th_drive.join(),th_ram.join()
    elif th_cpu != None and th_gpu != None and th_drive == None and th_ram != None:
        th_cpu.start(), th_gpu.start(), th_ram.start()
        th_cpu.join(),th_gpu.join(), th_ram.join()
    elif th_cpu != None and th_gpu != None and th_drive != None and th_ram == None:
        th_cpu.start(), th_gpu.start(), th_drive.start()
        th_cpu.join(),th_gpu.join(),th_drive.join()
    elif th_cpu == None and th_gpu == None and th_drive != None and th_ram != None:
        th_drive.start(), th_ram.start()
        th_drive.join(),th_ram.join()
    elif th_cpu == None and th_gpu != None and th_drive == None and th_ram != None:
        th_gpu.start(), th_ram.start()
        th_gpu.join(), th_ram.join()
    elif th_cpu == None and th_gpu != None and th_drive != None and th_ram == None:
        th_gpu.start(), th_drive.start()
        th_gpu.join(),th_drive.join()
    elif th_cpu != None and th_gpu == None and th_drive == None and th_ram != None:
        th_cpu.start(), th_ram.start()
        th_cpu.join(), th_ram.join()
    elif th_cpu != None and th_gpu == None and th_drive != None and th_ram == None:
        th_cpu.start(), th_drive.start()
        th_cpu.join(),th_drive.join()
    elif th_cpu != None and th_gpu != None and th_drive == None and th_ram == None:
        th_cpu.start(), th_gpu.start()
        th_cpu.join(),th_gpu.join()
    elif th_cpu != None and th_gpu == None and th_drive == None and th_ram == None:
        th_cpu.start()
        th_cpu.join()
    elif th_cpu == None and th_gpu != None and th_drive == None and th_ram == None:
        th_gpu.start()
        th_gpu.join()
    elif th_cpu == None and th_gpu == None and th_drive != None and th_ram == None:
        th_drive.start()
        th_drive.join()
    elif th_cpu == None and th_gpu == None and th_drive == None and th_ram != None:
        th_ram.start()
        th_ram.join()
    
    print("all Finish")
    os.rmdir("tmp")
    