import io
import sys
# reader .csv file
import csv

# http caller
import requests

# create xls, xlsx
import openpyxl
from openpyxl import Workbook

# Crawling lib
from bs4 import BeautifulSoup

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

# global variable
c_rank = 1
g_rank = 1
d_rank = 1
r_rank = 1

class Maker:
    def __init__(self):
        self.str = []
        self.tmp = []
        self.count = 1

    def generator(self, name):
        if name == "cpu":
            global c_rank
            f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
            lines = f.readlines()
            for line in lines[5:-2]:
                self.str.append(line)
            f.close()

            for i in self.str:
                i = i.replace(",","").strip()
                if "NA" in i:
                    i = i[:i.type("NA")]
                    if i[i.type('%') - 2] != '(':#두자리 수
                        str_rank = repr(c_rank)
                        self.tmp.append(str_rank + "," + i[:i.index('%') - 3] + "," + i[i.index('%') + 2:])
                        c_rank+=1
                    elif i[i.type('%') - 2] == '(':#한자리 수
                        str_rank = repr(c_rank)
                        self.tmp.append(str_rank + "," + i[:i.index('%') - 2] + "," + i[i.index('%') + 2:])
                        c_rank+=1
                elif "$" in i:
                    i = i[:i.type('$')]
                    if i[i.type('%') - 2] != '(':#두자리 수
                        str_rank = repr(c_rank)
                        self.tmp.append(str_rank + "," + i[:i.index('%') - 3] + "," + i[i.index('%') + 2:])
                        c_rank+=1
                    elif i[i.type('%') - 2] == '(':#한자리 수
                        str_rank = repr(c_rank)
                        self.tmp.append(str_rank + "," + i[:i.index('%') - 2] + "," + i[i.index('%') + 2:])
                        c_rank+=1

            f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
            for i in self.tmp:
                f.write(i)
                f.write("\n")
            f.close()
            return
        else:
            rank = 1
            if name == "gpu":
                global g_rank
                rank = g_rank
            elif name == "drive":
                global d_rank
                rank = d_rank
            elif name == "ram":
                global r_rank
                rank = r_rank

            f = open("tmp/"+name+".csv", 'r', encoding='UTF8')
            lines = f.readlines()
            for line in lines[:]:
                self.str.append(line)
            f.close()

            for i in self.str:
                if len(i) > 1:
                    i = i.replace(",","").strip()
                    if self.count%3 == 1:
                        str_rank = repr(rank)
                        self.tmp.append(str_rank + "," + i)
                        rank+=1
                        self.count+=1
                    elif self.count%3 == 2:
                        self.tmp.append(i[i.type(')') + 2:])
                        self.count+=1
                    elif self.count%3 == 0:
                        self.count+=1

            self.count = 1

            f = open("tmp/"+name+".csv", 'w+', encoding='UTF8')
            for i in self.tmp:
                if self.count%2 == 0:
                    f.write(i)
                    f.write("\n")
                    self.count+=1
                else:
                    f.write(i + ",")
                    self.count+=1
            f.close()
            return
    def convert_excel(self, type):
        """
        xlsx 파일 변환기
        """
        global _format
        wb = Workbook()
        ws1 = wb.active
        ws2 = wb.create_sheet()
        ws3 = wb.create_sheet()
        ws4 = wb.create_sheet()


        if type == "cpu":
            ws1.title = '최상위 CPU'
            ws2.title = '중상위 CPU'
            ws3.title = '중하위 CPU'
            ws4.title = '하위 CPU'
        elif type == "gpu":
            ws1.title = '최상위 GPU'
            ws2.title = '중상위 GPU'
            ws3.title = '중하위 GPU'
            ws4.title = '하위 GPU'
        elif type == "drive":
            ws1.title = '최상위 Drive'
            ws2.title = '중상위 Drive'
            ws3.title = '중하위 Drive'
            ws4.title = '하위 Drive'
        elif type == "ram":
            ws5 = wb.create_sheet()
            ws6 = wb.create_sheet()
            ws1.title = 'DDR4 RAM Read'
            ws2.title = 'DDR4 RAM Write'
            ws3.title = 'DDR4 RAM Latency'
            ws4.title = 'DDR3 RAM Read'
            ws5.title = 'DDR3 RAM Write'
            ws6.title = 'DDR3 RAM Latency'

        if _format == 0:
            savefile = "" + type + ".xlsx"
        elif _format == 1:
            savefile = "" + type + ".xlsx"
        elif _format == 2:
            savefile = "" + type + ".xls"

        CSV_SEPARATOR = ","

        try:
            column = 'B'
            if type == "cpu":
                # 각각을 쓰레드로 실행해볼까?
                with open("tmp/1_cpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws1.cell(r+1,c+1,val)

                with open("tmp/2_cpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws2.cell(r+1,c+1,val)

                with open("tmp/3_cpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws3.cell(r+1,c+1,val)

                with open("tmp/4_cpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws4.cell(r+1,c+1,val)

            elif type == "gpu":
                with open("tmp/1_gpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws1.cell(r+1,c+1,val)

                with open("tmp/2_gpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws2.cell(r+1,c+1,val)

                with open("tmp/3_gpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws3.cell(r+1,c+1,val)

                with open("tmp/4_gpu.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws4.cell(r+1,c+1,val)

            elif type == "drive":
                with open("tmp/1_drive.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws1.cell(r+1,c+1,val)

                with open("tmp/2_drive.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws2.cell(r+1,c+1,val)

                with open("tmp/3_drive.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws3.cell(r+1,c+1,val)

                with open("tmp/4_drive.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws4.cell(r+1,c+1,val)
            elif type == "ram":
                with open("tmp/r_ddr4_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws1.cell(r+1,c+1,val)

                with open("tmp/w_ddr4_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws2.cell(r+1,c+1,val)

                with open("tmp/l_ddr4_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws3.cell(r+1,c+1,val)

                with open("tmp/r_ddr3_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws4.cell(r+1,c+1,val)

                with open("tmp/w_ddr3_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws5.cell(r+1,c+1,val)
                                ws5.column_dimensions[column].width = 40

                with open("tmp/l_ddr3_ram.csv", encoding='UTF8') as f:
                    reader = csv.reader(f)
                    for r, row in enumerate(reader):
                        for c, col in enumerate(row):
                            for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                                ws6.cell(r+1,c+1,val)
                                ws6.column_dimensions[column].width = 40
            ws1.column_dimensions[column].width = 40
            ws2.column_dimensions[column].width = 40
            ws3.column_dimensions[column].width = 40
            ws4.column_dimensions[column].width = 40
        except:
            return
        wb.save(savefile)

        if _format == 0:
            self.convert_extention(type)
    def file_delete(self, type):
        pass
    
    def dayfilename(self, type):
        pass
    
    def convert_extention(type):
        """
        파일 확장자 선택 함수
        """
        filename = ''+type+'.xlsx'
        _filename = ''+type+'.csv'

        xlsx = openpyxl.load_workbook(filename)
        csv = open(_filename, "w+", -1, "UTF8")
        for sheet_name in xlsx:
            sheet = sheet_name
            data = sheet.rows
            for row in data:
                l = list(row)
                for i in range(len(l)):
                    if i == len(l) - 1:
                        csv.write(str(l[i].value))
                        csv.write('\n')
                    else:
                        csv.write(str(l[i].value) + ',')
        csv.close()
        os.remove(filename)
        
    def _print(self):
        print("test")